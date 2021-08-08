import os
from copy import copy

import openpyxl;

# Excel报表处理
from py_config import ConfigFactory
from py_logging import LoggerFactory


class ExcelHandle:
    def __init__(self, config, logger):
        self.config = config
        self.logger = logger
        self.template = config.get('data', 'template')
        self.output = config.get('data', 'output')
        # 数据表的表头
        self.data_header = ['序号', '统一编号', '原始编号', '取样量 (g/ml)', '分析结果\n ω(Au)/10-6', '分析结果\n ω(Au)/10-6',
                            '吸光度Abs', '备注']
        # 数据表
        self.data_list = [
            ['1', '标样', 'SG102', '30g', '', '', '', ''],
            ['2', '2020A-28055', 'SG101', '30g', '', '', '', ''],
            ['3', '2020A-28056', 'SG102', '30g', '', '', '', ''],
            ['4', '2020A-28057', 'SG103', '30g', '', '', '', ''],
            ['5', '2020A-28058', 'SG104', '30g', '', '', '', ''],
            ['6', '2020A-28059', 'SG105', '30g', '', '', '', ''],
            ['7', '2020A-28060', 'SG106', '30g', '', '', '', ''],
            ['8', '2020A-28060', 'SG107', '30g', '', '', '', ''],
            ['9', '2020A-28060', 'SG108', '30g', '', '', '', ''],
            ['10', '2020A-28061', 'SG109', '30g', '', '', '', ''],
            ['11', '2020A-28062', 'SG110', '30g', '', '', '', ''],
            ['12', '2020A-28063', 'SG111', '30g', '', '', '', ''],
        ]

        self.data_tail = ['', '称样人员：', '', '检测人员：', '', '', '校核人员：', '']

    # 带样式拷贝excel
    def copyfile_handler(self):
        # 打开templatebook和templatesheet
        template_workbook = openpyxl.load_workbook(self.template)
        template_worksheet = template_workbook.worksheets[0]

        # 获取template的最大行列数
        template_max_row = template_worksheet.max_row
        template_max_column = template_worksheet.max_column

        # 如果output存在
        if (os.path.exists(self.output)):
            output_workbook = openpyxl.load_workbook(self.output)
            output_worksheet = output_workbook.get_sheet_by_name('Sheet1', )
            # 获取output最大行号
            start_row_number = output_worksheet.max_row
            print('====begin====')
            print(start_row_number)
            print('====end====')
        else:
            # 创建输出workbook和worksheet
            output_workbook = openpyxl.Workbook()
            output_workbook.save(self.output)
            output_worksheet = output_workbook.create_sheet()
            # 获取output最大行号
            start_row_number = output_worksheet.max_row - 1

        # 带样式拷贝单元格
        for i in range(1, template_max_row + 1):
            for j in range(1, template_max_column + 1):
                source_cell = template_worksheet.cell(row=i, column=j)
                target_cell = output_worksheet.cell(row=i + start_row_number, column=j)
                target_cell.value = source_cell.value
                # 设置数据样式
                target_cell.data_type = source_cell.data_type
                target_cell.fill = copy(source_cell.fill)
                target_cell.border = copy(source_cell.border)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.alignment = copy(source_cell.alignment)
                target_cell.font = copy(source_cell.font)
        # 合并单元格
        # cell_rangs = ['A1:H1', 'A2:B2', 'D2:H2', 'G2:H2', 'A3:B3', 'A4:B4', 'G4:H4', 'A5:B5', 'E5:H5', 'A6:A7', 'B6:B7',
        #               'C6:C7', 'D6:D7', 'E6:F6', 'E7:F7', 'H6:H7']
        print('====begin====')
        print('start_row_number:%s' % start_row_number)
        print('A' + str(start_row_number + 1) + ':H' + str(start_row_number + 1))
        print('====end====')
        cell_rangs = ['A' + str(start_row_number + 1) + ':H' + str(start_row_number + 1),
                      'A' + str(start_row_number + 1) + ':B' + str(start_row_number + 1),
                      'D' + str(start_row_number + 2) + ':H' + str(start_row_number + 2),
                      'G' + str(start_row_number + 2) + ':H' + str(start_row_number + 2),
                      'A' + str(start_row_number + 3) + ':B' + str(start_row_number + 3),
                      'A' + str(start_row_number + 4) + ':B' + str(start_row_number + 4),
                      'G' + str(start_row_number + 4) + ':H' + str(start_row_number + 4),
                      'A' + str(start_row_number + 5) + ':B' + str(start_row_number + 5),
                      'E' + str(start_row_number + 5) + ':H' + str(start_row_number + 5),
                      'A' + str(start_row_number + 6) + ':A' + str(start_row_number + 7),
                      'B' + str(start_row_number + 6) + ':B' + str(start_row_number + 7),
                      'C' + str(start_row_number + 6) + ':C' + str(start_row_number + 7),
                      'D' + str(start_row_number + 6) + ':D' + str(start_row_number + 7),
                      'E' + str(start_row_number + 6) + ':F' + str(start_row_number + 6),
                      'E' + str(start_row_number + 7) + ':F' + str(start_row_number + 7),
                      'H' + str(start_row_number + 6) + ':H' + str(start_row_number + 7)]
        for cell_rang in cell_rangs:
            output_worksheet.merge_cells(cell_rang)

        # saving the destination excel file
        output_workbook.save(self.output)
        return output_workbook, output_worksheet

    # 数据替换处理函数
    def replace_handler(self, new_workbook, new_worksheet, replacement: dict):
        # sheet数据历遍
        for i in range(1, new_worksheet.max_row + 1):
            for j in range(1, new_worksheet.max_column + 1):
                cell_value = str(new_worksheet.cell(i, j).value)

                for key in replacement.keys():
                    if (key in cell_value):
                        new_worksheet.cell(i, j).value = cell_value.replace(key, str(replacement.get(key)))
        new_workbook.save(self.output)
        self.logger.info('The output file is %s' % self.output)
        return new_workbook, new_worksheet

    # excel文件定位写入处理
    def inject_handler(self, new_workbook, new_worksheet, data):
        data.append(self.data_tail)
        self.logger.info(data)
        start_row_num = new_worksheet.max_row
        self.logger.info(start_row_num)
        # sheet数据历遍
        for i in range(len(data)):
            for j in range(len(data[i])):
                new_worksheet.cell(i + 1 + start_row_num, j + 1).value = data[i][j]
        # print(i)
        # for k in range(len(self.data_tail)):
        #     new_worksheet.cell(i + 2 + start_row_num, k + 1).value = self.data_tail[k]
        new_workbook.save(self.output)
        self.logger.info('The output file is %s' % self.output)


if __name__ == '__main__':
    # 设置配置文件和日志
    config = ConfigFactory(config_file='py_jfk.ini').get_config()
    logger = LoggerFactory(config_factory=config).get_logger()
    excel_handle = ExcelHandle(config=config, logger=logger)
    book, sheet = excel_handle.copyfile_handler()
    # replacement = {'aa': 'aa123', 'bb': 'bb123', 'cc': 'cc12345'}
    # excel_handle.replace_handler(replacement)
    excel_handle.inject_handler(book, sheet, excel_handle.data_list)
